import json
import logging
import os
from datetime import datetime
from typing import Dict, Any, Optional

import pandas as pd
from PIL import Image as PILImage, ImageDraw, ImageFont
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak, Table, TableStyle


class ReportGenerator:

  def __init__(self, project) -> None:
    self.project = project
    self.output_dir = os.path.join(settings.MEDIA_ROOT, 'reports', str(project.id))
    os.makedirs(self.output_dir, exist_ok=True)

    try:
      self.font_path = os.path.join(settings.BASE_DIR, 'core/static/fonts/arial.ttf')
      self.font_main = ImageFont.truetype(self.font_path, 28)
      self.font_ticks = ImageFont.truetype(self.font_path, 24)
    except IOError:
      self.font_path = None
      self.font_main = ImageFont.load_default()
      self.font_ticks = ImageFont.load_default()

  def _create_grid_image(self, annotated_img_path: str, image_model, output_path: str) -> bool:
    try:
      source_img = PILImage.open(annotated_img_path).convert("RGBA")
      img_width_px, img_height_px = source_img.size
      scale = getattr(image_model, 'scale_factor', None)

      if not scale:
        source_img.convert('RGB').save(output_path, "PNG")
        return True

      real_width_cm = img_width_px / scale
      real_height_cm = img_height_px / scale
      margin = {'top': 60, 'bottom': 80, 'left': 90, 'right': 40}

      canvas_width = int(img_width_px + margin['left'] + margin['right'])
      canvas_height = int(img_height_px + margin['top'] + margin['bottom'])
      grid_canvas = PILImage.new('RGB', (canvas_width, canvas_height), 'white')
      draw = ImageDraw.Draw(grid_canvas)

      grid_color = (230, 230, 230)
      for i in range(0, int(real_width_cm) + 2, 2):
        x = int(margin['left'] + i * scale)
        draw.line([(x, margin['top']), (x, margin['top'] + img_height_px)],
                  fill=grid_color, width=1)

      for i in range(0, int(real_height_cm) + 2, 2):
        y = int(margin['top'] + i * scale)
        draw.line([(margin['left'], y), (margin['left'] + img_width_px, y)],
                  fill=grid_color, width=1)

      grid_canvas.paste(source_img, (margin['left'], margin['top']), source_img)

      axis_color = (60, 60, 60)
      draw.line([(margin['left'], margin['top']),
                 (margin['left'], margin['top'] + img_height_px)],
                fill=axis_color, width=2)

      for i in range(0, int(real_height_cm) + 2, 2):
        y = int(margin['top'] + i * scale)
        draw.line([(margin['left'] - 5, y), (margin['left'], y)],
                  fill=axis_color, width=2)
        draw.text((margin['left'] - 42, y - 8), str(i),
                  fill=axis_color, font=self.font_ticks)

      draw.line([(margin['left'], margin['top'] + img_height_px),
                 (margin['left'] + img_width_px, margin['top'] + img_height_px)],
                fill=axis_color, width=2)

      for i in range(0, int(real_width_cm) + 2, 2):
        x = int(margin['left'] + i * scale)
        draw.line([(x, margin['top'] + img_height_px),
                   (x, margin['top'] + img_height_px + 5)],
                  fill=axis_color, width=2)
        draw.text((x - 8, margin['top'] + img_height_px + 8), str(i),
                  fill=axis_color, font=self.font_ticks)

      draw.text((canvas_width / 2 - 50, canvas_height - 40),
                "Ancho (cm)", fill=axis_color, font=self.font_main)

      y_label_img = PILImage.new('L', (200, 50))
      y_draw = ImageDraw.Draw(y_label_img)
      y_draw.text((0, 0), "Alto (cm)", font=self.font_main, fill=255)
      rotated_label = y_label_img.rotate(90, expand=1)

      paste_x = 15
      paste_y = int(canvas_height / 2 - rotated_label.size[1] / 2)
      grid_canvas.paste(PILImage.new('RGB', rotated_label.size, 'white'),
                        (paste_x, paste_y), rotated_label)

      grid_canvas.save(output_path, "PNG")
      return True

    except Exception as e:
      logging.exception("Failed to create grid image: %s", e)
      if os.path.exists(annotated_img_path):
        import shutil
        shutil.copy(annotated_img_path, output_path)
      return False

  def generate_pdf_for_image(self, image, include_measurements=True) -> str:
    filename = f"reporte_imagen_{image.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(self.output_dir, filename)

    grid_image_path = os.path.join(self.output_dir, f"temp_grid_{image.id}.png")

    try:
      doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=0.7 * inch, leftMargin=0.7 * inch,
                              topMargin=0.7 * inch, bottomMargin=0.7 * inch)
      story = []
      styles = getSampleStyleSheet()

      title_style = ParagraphStyle('Title', parent=styles['h1'], alignment=TA_CENTER, fontSize=14, spaceAfter=12)
      heading_style = ParagraphStyle('Heading', parent=styles['h2'], fontSize=11, spaceBefore=6, spaceAfter=3,
                                     textColor=colors.HexColor('#2c3e50'))
      body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=9, spaceAfter=1, leading=11)
      bullet_style = ParagraphStyle('Bullet', parent=body_style, leftIndent=15)
      footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey,
                                    alignment=TA_LEFT, leading=9)

      story.append(Paragraph("Mediciones de Objetos - Referencia ArUco (5.0x5.0 cm)", title_style))

      annotated_path = image.image.path
      if image.calibration_data and image.calibration_data.get('annotated_url'):
        relative_path = image.calibration_data['annotated_url'].replace(settings.MEDIA_URL, '')
        full_annotated_path = os.path.join(settings.MEDIA_ROOT, relative_path)
        if os.path.exists(full_annotated_path):
          annotated_path = full_annotated_path

      if annotated_path and self._create_grid_image(annotated_path, image, grid_image_path):
        page_width, page_height = A4
        available_width = page_width - 1.4 * inch
        available_height = page_height - 1.4 * inch - 100

        img_pil = PILImage.open(grid_image_path)
        img_width, img_height = img_pil.size
        img_pil.close()

        aspect_ratio = img_width / img_height

        if img_width > available_width or img_height > available_height:
          width_constrained = available_width
          height_constrained_by_width = width_constrained / aspect_ratio

          height_constrained = available_height
          width_constrained_by_height = height_constrained * aspect_ratio

          if height_constrained_by_width <= available_height:
            final_width = width_constrained
            final_height = height_constrained_by_width
          else:
            final_width = width_constrained_by_height
            final_height = height_constrained
        else:
          final_width = img_width * 72 / 96
          final_height = img_height * 72 / 96

        if final_width > 4.5 * inch:
          final_width = 4.5 * inch
          final_height = final_width / aspect_ratio

        img_obj = Image(grid_image_path, width=final_width, height=final_height)
        story.append(img_obj)
        story.append(Spacer(1, 10))

      detected_shapes = image.shapes.filter(measurement__is_manual=False).distinct()
      if detected_shapes.exists():
        story.append(Paragraph("Reporte de Mediciones", heading_style))
        for i, shape in enumerate(detected_shapes, 1):
          props = shape.properties or {}
          shape_display_name = shape.get_shape_type_display().upper()
          story.append(Paragraph(f"<b>OBJETO {i}: {shape_display_name}</b>", body_style))

          bullet_points = []
          if shape.shape_type == 'rectangle':
            if props.get('w_cm'): bullet_points.append(f"Ancho: {props['w_cm']:.2f} cm")
            if props.get('h_cm'): bullet_points.append(f"Alto: {props['h_cm']:.2f} cm")
          elif shape.shape_type == 'circle':
            if props.get('radius_cm'): bullet_points.append(f"Radio: {props['radius_cm']:.2f} cm")
            if props.get('diameter_cm'): bullet_points.append(f"Diámetro: {props['diameter_cm']:.2f} cm")
          elif shape.shape_type == 'triangle':
            sides = props.get('sides_cm', [])
            if len(sides) >= 3:
              bullet_points.append(f"Lado A: {sides[0]:.2f} cm")
              bullet_points.append(f"Lado B: {sides[1]:.2f} cm")
              bullet_points.append(f"Lado C: {sides[2]:.2f} cm")
          elif shape.shape_type == 'polygon':
            if props.get('w_cm'): bullet_points.append(f"Ancho: {props['w_cm']:.2f} cm")
            if props.get('h_cm'): bullet_points.append(f"Alto: {props['h_cm']:.2f} cm")

          if props.get('area_cm2'): bullet_points.append(f"Área aproximada: {props['area_cm2']:.2f} cm²")
          if props.get('perimeter_cm'): bullet_points.append(f"Perímetro: {props['perimeter_cm']:.2f} cm")
          center = props.get('raw', {}).get('center_px')
          if center: bullet_points.append(f"Centro (px): ({int(center[0])}, {int(center[1])})")

          for point in bullet_points:
            story.append(Paragraph(f"&bull; {point}", bullet_style))
          story.append(Spacer(1, 6))

      manual_measurements = image.measurements.filter(is_manual=True).order_by('created_at')
      if manual_measurements.exists():
        story.append(Paragraph("Medidas Manuales", heading_style))
        for i, measurement in enumerate(manual_measurements, 1):
          story.append(Paragraph(f"&bull; Medida {i}: {measurement.value_display}", bullet_style))
        story.append(Spacer(1, 10))

      story.append(Paragraph("=" * 80, footer_style))
      scale_info = "No calibrada"
      if image.is_calibrated and image.scale_factor:
        scale_info = f"Escala de conversión: {image.scale_factor:.2f} píxeles/{image.calibration_unit}"

      story.append(Paragraph(f"Total de objetos detectados: {detected_shapes.count()}", footer_style))
      story.append(Paragraph(f"Reporte generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))

      doc.build(story)
    finally:
      if os.path.exists(grid_image_path):
        os.remove(grid_image_path)

    return filepath

  def _fmt(self, v: Any) -> Optional[float]:
    try:
      if v is None:
        return None
      return float(f"{float(v):.2f}")
    except Exception:
      return v

  def generate_excel_for_image(self, image) -> str:
    filename = f"datos_imagen_{image.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(self.output_dir, filename)

    with pd.ExcelWriter(path=filepath, engine='openpyxl') as writer:
      detected_objects_data = []
      detected_shapes = image.shapes.filter(measurement__is_manual=False).distinct()

      for i, shape in enumerate(detected_shapes, 1):
        props = shape.properties or {}
        raw_props = props.get('raw', {})
        center = raw_props.get('center_px', [None, None])
        sides = props.get('sides_cm', [])

        detected_objects_data.append({
          'ID Objeto': i,
          'Nombre': shape.name,
          'Tipo': shape.get_shape_type_display(),
          'Ancho (cm)': props.get('w_cm'),
          'Alto (cm)': props.get('h_cm'),
          'Área (cm²)': props.get('area_cm2'),
          'Perímetro (cm)': props.get('perimeter_cm'),
          'Lado A (cm)': sides[0] if len(sides) > 0 else None,
          'Lado B (cm)': sides[1] if len(sides) > 1 else None,
          'Lado C (cm)': sides[2] if len(sides) > 2 else None,
          'Radio (cm)': props.get('radius_cm'),
          'Diámetro (cm)': props.get('diameter_cm') or (
            props.get('radius_cm', 0) * 2 if props.get('radius_cm') else None
          ),
          'Centro X (px)': center[0],
          'Centro Y (px)': center[1],
          'Fecha Detección': shape.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })

      if detected_objects_data:
        pd.DataFrame(detected_objects_data).dropna(axis=1, how='all').to_excel(
          writer,
          sheet_name='Objetos Detectados',
          index=False
        )

      manual_measurements_data = []
      manual_measurements = image.measurements.filter(is_manual=True).order_by('created_at')

      for i, measurement in enumerate(manual_measurements, 1):
        manual_measurements_data.append({
          'ID Medida': i,
          'Nombre': measurement.name,
          'Valor': measurement.value_real,
          'Unidad': measurement.unit,
          'Fecha Creación': measurement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })

      if manual_measurements_data:
        pd.DataFrame(manual_measurements_data).to_excel(
          writer,
          sheet_name='Medidas Manuales',
          index=False
        )
    return filepath

  def generate_pdf_report(self, include_images: bool = True, include_measurements: bool = True,
                          include_shapes: bool = True) -> str:
    filename = f"reporte_completo_{self.project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    filepath = os.path.join(self.output_dir, filename)
    doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=12)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceAfter=8)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9, leading=11)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, leading=10)

    total_images = self.project.images.count()
    total_measurements = sum(img.measurements.count() for img in self.project.images.all())

    story = [Paragraph(f"Reporte de Proyecto: {self.project.name}", title_style)]
    meta_lines = [
      f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
      f"Imágenes en proyecto: {total_images}",
      f"Total mediciones: {total_measurements}"
    ]
    for ml in meta_lines:
      story.append(Paragraph(ml, small_style))
    story.append(Spacer(1, 12))

    for image in self.project.images.all():
      story.append(Paragraph(f"Imagen: {image.name}", heading_style))
      if include_images and getattr(image, 'image', None):
        try:
          img_path = image.image.path
          img_obj = Image(img_path, width=4 * inch, height=3 * inch)
          story.append(img_obj)
          story.append(Spacer(1, 6))
        except Exception:
          pass

      auto_ms = image.measurements.filter(is_manual=False).order_by('created_at')
      if include_measurements:
        story.append(Paragraph('Mediciones Automáticas', styles['Heading4']))
        if auto_ms.exists():
          data = [['#', 'Nombre', 'Tipo', 'Valor', 'Unidad', 'Forma Asociada', 'Fecha']]
          for i, m in enumerate(auto_ms, 1):
            val = m.value_real if m.value_real is not None else m.value_pixels
            try:
              val_fmt = f"{float(val):.2f}"
            except Exception:
              val_fmt = str(val)
            shape_name = m.shape.name if getattr(m, 'shape', None) else ''
            data.append([str(i), m.name, m.measurement_type or '', val_fmt, m.unit or '', shape_name,
                         m.created_at.strftime('%Y-%m-%d %H:%M:%S')])
          table = Table(data, colWidths=[24, 120, 60, 60, 40, 90, 80])
          table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
          ]))
          story.append(table)
          story.append(Spacer(1, 8))
        else:
          story.append(Paragraph('No se encontraron mediciones automáticas para esta imagen.', normal_style))
          story.append(Spacer(1, 6))

        manual_ms = image.measurements.filter(is_manual=True).order_by('created_at')
        story.append(Paragraph('Mediciones Manuales', styles['Heading4']))
        if manual_ms.exists():
          data = [['#', 'Nombre', 'Valor', 'Unidad', 'Fecha']]
          for i, m in enumerate(manual_ms, 1):
            val = m.value_real if m.value_real is not None else m.value_pixels
            try:
              val_fmt = f"{float(val):.2f}"
            except Exception:
              val_fmt = str(val)
            data.append([str(i), m.name, val_fmt, m.unit or '', m.created_at.strftime('%Y-%m-%d %H:%M:%S')])
          table = Table(data, colWidths=[24, 200, 80, 50, 100])
          table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
          ]))
          story.append(table)
          story.append(Spacer(1, 12))
        else:
          story.append(Paragraph('No se encontraron mediciones manuales para esta imagen.', normal_style))
          story.append(Spacer(1, 12))

      story.append(PageBreak())

    story.append(Paragraph(f"Fin del reporte - {self.project.name}", small_style))
    doc.build(story)
    return filepath

  def generate_excel_report(self, include_images: bool = True, include_measurements: bool = True,
                            include_shapes: bool = True) -> str:
    filename = f"datos_completos_{self.project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(self.output_dir, filename)
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path=filepath, engine='openpyxl') as writer:
      project_data = {'Clave': ['Nombre del Proyecto'], 'Valor': [self.project.name]}
      pd.DataFrame(project_data).to_excel(writer, sheet_name='Resumen Proyecto', index=False)

      for image in self.project.images.all():
        base_name = image.name[:20].replace('/', '_')
        if include_shapes:
          shapes_rows = []
          for i, shape in enumerate(image.shapes.all(), 1):
            props = shape.properties or {}
            sides = props.get('sides_cm', [])
            shapes_rows.append({
              'ID Objeto': i,
              'Nombre': shape.name,
              'Tipo': shape.get_shape_type_display(),
              'Ancho (cm)': self._fmt(props.get('w_cm')),
              'Alto (cm)': self._fmt(props.get('h_cm')),
              'Área (cm²)': self._fmt(props.get('area_cm2')),
              'Perímetro (cm)': self._fmt(props.get('perimeter_cm')),
              'Lado A (cm)': self._fmt(sides[0] if len(sides) > 0 else None),
              'Lado B (cm)': self._fmt(sides[1] if len(sides) > 1 else None),
              'Lado C (cm)': self._fmt(sides[2] if len(sides) > 2 else None),
              'Radio (cm)': self._fmt(props.get('radius_cm')),
              'Diámetro (cm)': self._fmt(
                props.get('diameter_cm') or (props.get('radius_cm', 0) * 2 if props.get('radius_cm') else None)),
              'Fecha Detección': shape.created_at.strftime('%Y-%m-%d %H:%M:%S') if getattr(shape, 'created_at',
                                                                                           None) else ''
            })
          if shapes_rows:
            df_shapes = pd.DataFrame(shapes_rows)
            sheet_name = f"{base_name}_Formas"[:31]
            df_shapes.to_excel(writer, sheet_name=sheet_name, index=False)

        if include_measurements:
          meas_rows = []
          for m in image.measurements.all():
            meas_rows.append({
              'Nombre': m.name,
              'Tipo': m.measurement_type,
              'Valor Real': self._fmt(m.value_real),
              'Unidad': m.unit,
              'Valor (px)': self._fmt(m.value_pixels),
              'Es Manual': bool(m.is_manual),
              'Fecha': m.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
          if meas_rows:
            df_meas = pd.DataFrame(meas_rows)
            sheet_name = f"{base_name}_Mediciones"[:31]
            df_meas.to_excel(writer, sheet_name=sheet_name, index=False)

    try:
      import openpyxl
      wb = openpyxl.load_workbook(filepath)
      for ws in wb.worksheets:
        for i, col in enumerate(ws.columns, 1):
          max_length = 0
          column = get_column_letter(i)
          for cell in col:
            try:
              value = str(cell.value) if cell.value is not None else ''
              if len(value) > max_length:
                max_length = len(value)
            except Exception:
              pass
          adjusted_width = (max_length + 2) * 1.1
          ws.column_dimensions[column].width = min(50, adjusted_width)

        for row in ws.iter_rows(min_row=2):
          for cell in row:
            try:
              if isinstance(cell.value, float):
                cell.number_format = '0.00'
            except Exception:
              pass

      wb.save(filepath)
    except Exception:
      pass

    return filepath

  def generate_json_report(self, include_images: bool = True, include_measurements: bool = True,
                           include_shapes: bool = True) -> str:
    filename = f"datos_completos_{self.project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.json"
    filepath = os.path.join(self.output_dir, filename)
    report_data = {
      'project': {
        'id': str(self.project.id),
        'name': self.project.name,
        'images': []
      }
    }

    for image in self.project.images.all():
      img = {
        'id': str(image.id),
        'name': image.name
      }
      if include_images and image.image:
        img['image_url'] = image.image.url
      if include_shapes:
        img['shapes'] = []
        for s in image.shapes.all():
          img['shapes'].append({
            'id': str(s.id),
            'name': s.name,
            'type': s.shape_type,
            'points': s.points,
            'properties': s.properties
          })
      if include_measurements:
        img['measurements'] = []
        for m in image.measurements.all():
          img['measurements'].append({
            'id': str(m.id),
            'name': m.name,
            'type': m.measurement_type,
            'value_pixels': m.value_pixels,
            'value_real': m.value_real,
            'unit': m.unit,
            'created_at': m.created_at.isoformat()
          })

      report_data['project']['images'].append(img)

    with open(filepath, 'w', encoding='utf-8') as f:
      json.dump(report_data, f, indent=2, ensure_ascii=False)
    return filepath


class QuickReportGenerator:

  @staticmethod
  def generate_measurement_summary(project) -> Dict[str, Any]:
    summary = {
      'project_name': project.name,
      'total_images': project.images.count(),
      'total_measurements': 0,
      'measurement_types': {},
      'calibrated_images': 0,
      'units_used': set(),
      'images_summary': []
    }
    for image in project.images.all():
      if image.is_calibrated:
        summary['calibrated_images'] += 1
        summary['units_used'].add(image.calibration_unit)

      image_measurements = image.shapes.count()
      summary['total_measurements'] += image_measurements
      for measurement in image.measurements.all():
        mtype = measurement.measurement_type
        summary['measurement_types'][mtype] = summary['measurement_types'].get(mtype, 0) + 1
      summary['images_summary'].append({
        'name': image.name,
        'measurements_count': image_measurements,
        'shapes_count': image.shapes.count(),
        'is_calibrated': image.is_calibrated
      })
    summary['units_used'] = list(summary['units_used'])
    return summary
